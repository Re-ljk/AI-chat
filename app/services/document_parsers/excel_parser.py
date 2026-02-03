"""
    @project: aihub
    @Author: jiangkuanli
    @file: excel_parser
    @date: 2026/2/2
    @desc: Excel文档解析器
"""

from typing import Dict, Any, List
import base64
from .base_parser import BaseDocumentParser


class ExcelParser(BaseDocumentParser):
    """Excel文档解析器"""
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """解析Excel文档
        
        Args:
            file_path: Excel文档路径
            
        Returns:
            包含文档内容和元数据的字典
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            
            df = pd.read_excel(file_path)
            content = []
            images = []
            
            for column in df.columns:
                column_data = df[column].dropna().astype(str).tolist()
                content.append(f"{column}:\n" + "\n".join(column_data))
                content.append("\n")
            
            full_text = '\n'.join(content)
            cleaned_text = self.clean_text(full_text)
            
            image_count = self._extract_images(file_path, images)
            
            metadata = self.extract_metadata(cleaned_text)
            metadata.update({
                'sheet_count': len(df.sheet_names),
                'column_count': len(df.columns),
                'row_count': len(df),
                'image_count': image_count,
                'has_images': image_count > 0,
                'file_type': 'excel'
            })
            
            return {
                'content': cleaned_text,
                'metadata': metadata,
                'images': images,
                'success': True
            }
            
        except ImportError:
            return {
                'content': '',
                'metadata': {},
                'images': [],
                'success': False,
                'error': 'pandas或openpyxl库未安装，请运行: pip install pandas openpyxl'
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {},
                'images': [],
                'success': False,
                'error': f'解析Excel文档失败: {str(e)}'
            }
    
    def _extract_images(self, file_path: str, images: List[Dict[str, Any]]) -> int:
        """提取Excel文档中的图片
        
        Args:
            file_path: 文件路径
            images: 图片列表
            
        Returns:
            图片数量
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            image_count = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                if hasattr(ws, '_images'):
                    for image in ws._images:
                        try:
                            image_data = image._data()
                            if image_data:
                                image_info = {
                                    'index': image_count,
                                    'sheet': sheet_name,
                                    'extension': image.format.lower(),
                                    'size': len(image_data),
                                    'data': base64.b64encode(image_data).decode('utf-8')
                                }
                                images.append(image_info)
                                image_count += 1
                        except Exception as e:
                            print(f"提取图片 {image_count} 时出错: {str(e)}")
                            continue
            
            return image_count
        except Exception as e:
            print(f"提取Excel图片时出错: {str(e)}")
            return 0
    
    def get_supported_extensions(self) -> list:
        """获取支持的文件扩展名"""
        return ['.xlsx', '.xls', '.csv']
